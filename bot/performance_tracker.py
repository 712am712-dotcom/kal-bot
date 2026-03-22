"""
performance_tracker.py — Tracks Kal's prediction accuracy and live P&L.

Records every non-SKIP recommendation (research and live) to a persistent CSV.
Polls the Kalshi API at each scan cycle to detect market resolution, updates
WIN/LOSS status, computes P&L, and rebuilds a colour-formatted XLSX workbook.

Architecture
------------
- CSV  is the source of truth.  Written atomically on every change.
- XLSX is a derived view rebuilt from CSV after every update.
- asyncio.Lock protects file operations; Kalshi API calls happen outside it.

P&L accounting
--------------
- BUY YES at price P:  win → +contracts × (1 − P)    loss → −bet_dollars
- BUY NO  at price P:  win → +contracts × P            loss → −bet_dollars
  (where P = kalshi_yes_price, stored as a fraction 0–1)
- Research rows always record P&L = $0 (no real money deployed).
- Running Total P&L counts LIVE-mode rows only, in chronological order.
"""
from __future__ import annotations

import asyncio
import csv
import json
from datetime import date, datetime, timezone
from pathlib import Path
from typing import TYPE_CHECKING

import structlog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from claude_client import MarketAnalysis
    from kalshi_client import KalshiClient

log = structlog.get_logger(__name__)

# ── File paths ─────────────────────────────────────────────────────────────────
_HERE     = Path(__file__).parent
CSV_PATH  = _HERE / "kal_performance.csv"
XLSX_PATH = _HERE / "kal_performance.xlsx"

# ── Column schema ──────────────────────────────────────────────────────────────
# Display columns — appear in the XLSX (in this exact order)
DISPLAY_COLS: list[str] = [
    "Date",
    "Time",
    "Coin",
    "Market Title",
    "Timeframe",
    "Recommendation",
    "Confidence %",
    "Edge %",
    "Kalshi Price at Entry",
    "Kal Probability",
    "Amount Wagered ($)",
    "Result",
    "Profit/Loss ($)",
    "Running Total P&L ($)",
    "Mode",
]

# Internal columns — stored in CSV for resolution logic, hidden in XLSX
_INTERNAL_COLS: list[str] = ["_ticker", "_side", "_contracts", "_yes_price_cents"]
ALL_COLS: list[str] = DISPLAY_COLS + _INTERNAL_COLS

# Columns that hold decimal probabilities (displayed as percentages in XLSX)
_PCT_COLS   = {"Confidence %", "Edge %", "Kalshi Price at Entry", "Kal Probability"}
# Columns that hold dollar amounts (formatted as currency in XLSX)
_MONEY_COLS = {"Amount Wagered ($)", "Profit/Loss ($)", "Running Total P&L ($)"}

# ── Excel style constants ──────────────────────────────────────────────────────
_THIN        = Side(style="thin", color="AAAAAA")
_BORDER      = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_HEADER_FILL  = PatternFill("solid", fgColor="1F3864")   # dark navy blue
_HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_WIN_FILL     = PatternFill("solid", fgColor="C6EFCE")   # light green
_LOSS_FILL    = PatternFill("solid", fgColor="FFC7CE")   # light red
_PENDING_FILL = PatternFill("solid", fgColor="FFEB9C")   # light yellow
_SUMMARY_FILL = PatternFill("solid", fgColor="BDD7EE")   # light steel blue
_SUMMARY_FONT = Font(bold=True, name="Calibri", size=10)
_BODY_FONT    = Font(name="Calibri", size=10)
_CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=False)
_LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=False)


# ── Helpers ────────────────────────────────────────────────────────────────────

def _blank_row() -> dict:
    return {c: "" for c in ALL_COLS}


def _pct_display(raw: str) -> str:
    """Format a raw decimal string as a percentage string for length-measurement."""
    try:
        return f"{float(raw):.1%}"
    except (ValueError, TypeError):
        return str(raw)


def _money_display(raw: str) -> str:
    """Format a raw float string as a currency string for length-measurement."""
    try:
        return f"${float(raw):,.2f}"
    except (ValueError, TypeError):
        return str(raw)


# ── Core class ─────────────────────────────────────────────────────────────────

class PerformanceTracker:
    """Async-safe tracker: CSV persistence + XLSX reporting."""

    def __init__(
        self,
        csv_path:  Path = CSV_PATH,
        xlsx_path: Path = XLSX_PATH,
    ) -> None:
        self._csv  = csv_path
        self._xlsx = xlsx_path
        self._lock = asyncio.Lock()
        self._init_csv()

    # ── CSV I/O ────────────────────────────────────────────────────────────────

    def _init_csv(self) -> None:
        if not self._csv.exists():
            with open(self._csv, "w", newline="", encoding="utf-8") as fh:
                csv.DictWriter(fh, fieldnames=ALL_COLS).writeheader()
            log.info("performance_csv_created", path=str(self._csv))

    def _read(self) -> list[dict]:
        with open(self._csv, "r", newline="", encoding="utf-8") as fh:
            return list(csv.DictReader(fh))

    def _write(self, rows: list[dict]) -> None:
        with open(self._csv, "w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=ALL_COLS)
            w.writeheader()
            for row in rows:
                w.writerow({c: row.get(c, "") for c in ALL_COLS})

    # ── Public API ─────────────────────────────────────────────────────────────

    async def record(
        self,
        analysis:    "MarketAnalysis",
        side:        str,    # "yes" | "no"
        contracts:   int,    # 0 for research
        bet_dollars: float,  # 0.0 for research
        mode:        str,    # "research" | "live"
    ) -> None:
        """
        Append a PENDING row for a new decision.
        Research rows use contracts=0, bet_dollars=0.0.
        """
        now = datetime.now(timezone.utc)
        row = _blank_row()

        row["Date"]         = now.strftime("%Y-%m-%d")
        row["Time"]         = now.strftime("%H:%M:%S")
        row["Coin"]         = analysis.coin        or "—"
        row["Market Title"] = analysis.market_title
        row["Timeframe"]    = analysis.timeframe   or "—"
        row["Mode"]         = mode.upper()

        row["Recommendation"]        = analysis.recommendation.replace("_", " ")
        # Store as raw decimal so XLSX can apply percentage format
        row["Confidence %"]          = f"{analysis.confidence:.6f}"
        row["Edge %"]                = f"{analysis.edge:.6f}"
        row["Kalshi Price at Entry"] = f"{analysis.kalshi_yes_price:.6f}"
        row["Kal Probability"]       = f"{analysis.yes_probability:.6f}"

        row["Amount Wagered ($)"]    = f"{bet_dollars:.2f}"
        row["Result"]                = "PENDING"
        row["Profit/Loss ($)"]       = "0.00"
        row["Running Total P&L ($)"] = ""

        # Internal (resolution)
        row["_ticker"]          = analysis.ticker
        row["_side"]            = side.lower()
        row["_contracts"]       = str(contracts)
        row["_yes_price_cents"] = str(round(analysis.kalshi_yes_price * 100))

        async with self._lock:
            rows = self._read()
            rows.append(row)
            self._write(rows)

        log.info(
            "performance_recorded",
            ticker=analysis.ticker,
            recommendation=analysis.recommendation,
            mode=mode,
            contracts=contracts,
            bet=f"${bet_dollars:.2f}",
        )

    async def resolve_pending(self, kalshi: "KalshiClient") -> list[dict]:
        """
        Poll Kalshi for each PENDING row. Update WIN/LOSS and P&L where resolved.
        Recomputes running totals for all LIVE rows after any change.
        Rebuilds XLSX if any rows were updated.
        Returns the list of newly-resolved row dicts.
        """
        # ── Step 1: snapshot pending rows (brief lock) ─────────────────────
        async with self._lock:
            snapshot = self._read()
        pending: list[tuple[int, dict]] = [
            (i, r) for i, r in enumerate(snapshot)
            if r.get("Result") == "PENDING"
        ]
        if not pending:
            return []

        # ── Step 2: poll Kalshi — NO lock held during I/O ──────────────────
        # Map: row_index → {"result": "yes"|"no", ...}
        resolutions: dict[int, str] = {}  # index → "yes" | "no"

        for idx, row in pending:
            ticker = row.get("_ticker", "").strip()
            if not ticker:
                continue
            try:
                market = await kalshi.get_market(ticker)
            except Exception as exc:
                log.warning("resolve_fetch_failed", ticker=ticker, error=str(exc))
                continue

            status = (market.get("status") or "").lower()
            result = (market.get("result") or "").lower()

            if status not in ("determined", "finalized", "closed"):
                continue
            if result not in ("yes", "no"):
                continue

            resolutions[idx] = result
            log.debug("market_resolved_detected", ticker=ticker, result=result, status=status)

        if not resolutions:
            return []

        # ── Step 3: apply updates atomically (re-read for any new appends) ─
        newly_resolved: list[dict] = []

        async with self._lock:
            rows = self._read()

            for idx, market_result in resolutions.items():
                if idx >= len(rows):
                    continue  # row was somehow removed (shouldn't happen)

                row  = rows[idx]
                side = row.get("_side", "").lower()
                won  = (
                    (side == "yes" and market_result == "yes") or
                    (side == "no"  and market_result == "no")
                )
                row["Result"] = "WIN" if won else "LOSS"

                # Compute P&L
                if row.get("Mode", "").upper() == "LIVE":
                    try:
                        contracts      = int(row.get("_contracts") or 0)
                        yes_p_cents    = int(row.get("_yes_price_cents") or 0)
                        bet            = float(row.get("Amount Wagered ($)") or 0)
                        yes_price      = yes_p_cents / 100.0

                        if side == "yes":
                            pnl = contracts * (1.0 - yes_price) if won else -bet
                        else:
                            no_price = 1.0 - yes_price
                            pnl      = contracts * (1.0 - no_price) if won else -bet
                    except (ValueError, TypeError):
                        pnl = 0.0
                    row["Profit/Loss ($)"] = f"{pnl:.2f}"
                else:
                    row["Profit/Loss ($)"] = "0.00"

                rows[idx] = row
                newly_resolved.append(dict(row))

                log.info(
                    "market_resolved",
                    ticker=row.get("_ticker"),
                    market_result=market_result,
                    our_side=side,
                    won=won,
                    pnl=row["Profit/Loss ($)"],
                    mode=row.get("Mode"),
                )

            # Recompute running P&L for all LIVE rows (chronological order)
            running = 0.0
            for r in rows:
                if r.get("Mode", "").upper() != "LIVE":
                    r["Running Total P&L ($)"] = ""
                    continue
                result_val = r.get("Result", "")
                if result_val in ("WIN", "LOSS"):
                    try:
                        running += float(r.get("Profit/Loss ($)") or 0)
                    except ValueError:
                        pass
                r["Running Total P&L ($)"] = f"{running:.2f}"

            self._write(rows)

        # ── Step 4: rebuild XLSX outside the lock ──────────────────────────
        self._rebuild_excel()
        return newly_resolved

    def rebuild_excel(self) -> None:
        """Public wrapper — call after bulk imports or manual CSV edits."""
        self._rebuild_excel()

    def _rebuild_excel(self) -> None:
        """Regenerate XLSX from current CSV data with full colour formatting."""
        rows = self._read()
        wb   = Workbook()
        ws   = wb.active
        ws.title = "Kal Performance"
        ws.freeze_panes = "A2"
        ws.sheet_properties.tabColor = "1F3864"

        # ── Header row ─────────────────────────────────────────────────────
        for c, name in enumerate(DISPLAY_COLS, start=1):
            cell           = ws.cell(row=1, column=c, value=name)
            cell.fill      = _HEADER_FILL
            cell.font      = _HEADER_FONT
            cell.alignment = _CENTER
            cell.border    = _BORDER

        # Track max display-string length per column for auto-sizing
        col_max: list[int] = [len(name) for name in DISPLAY_COLS]

        # ── Data rows ──────────────────────────────────────────────────────
        for r_idx, row in enumerate(rows, start=2):
            result = row.get("Result", "PENDING")
            fill   = (
                _WIN_FILL     if result == "WIN"  else
                _LOSS_FILL    if result == "LOSS" else
                _PENDING_FILL
            )

            for c_idx, name in enumerate(DISPLAY_COLS, start=1):
                raw = (row.get(name) or "").strip()

                # Convert to native Python type for proper Excel formatting
                if name in _PCT_COLS:
                    try:
                        value = float(raw)
                    except ValueError:
                        value = raw
                    display_str = _pct_display(raw)
                elif name in _MONEY_COLS:
                    try:
                        value = float(raw)
                    except ValueError:
                        value = raw
                    display_str = _money_display(raw)
                else:
                    value = raw
                    display_str = raw

                cell           = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.fill      = fill
                cell.font      = _BODY_FONT
                cell.border    = _BORDER
                cell.alignment = _LEFT if name == "Market Title" else _CENTER

                if name in _PCT_COLS:
                    cell.number_format = "0.0%"
                elif name in _MONEY_COLS:
                    cell.number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'

                col_max[c_idx - 1] = max(col_max[c_idx - 1], len(display_str))

        # ── Summary row ────────────────────────────────────────────────────
        if rows:
            summary_r   = len(rows) + 3     # one blank row gap

            all_closed  = [r for r in rows if r.get("Result") in ("WIN", "LOSS")]
            live_closed = [r for r in all_closed if r.get("Mode", "").upper() == "LIVE"]
            wins        = sum(1 for r in all_closed if r.get("Result") == "WIN")
            losses      = len(all_closed) - wins
            pending_cnt = sum(1 for r in rows if r.get("Result") == "PENDING")
            win_rate    = wins / len(all_closed) * 100 if all_closed else 0.0

            total_pnl = 0.0
            for r in live_closed:
                try:
                    total_pnl += float(r.get("Profit/Loss ($)") or 0)
                except ValueError:
                    pass

            # Build sparse dict of summary values keyed by column name
            summary: dict[str, object] = {
                "Date":                    "SUMMARY",
                "Coin":                    f"{len(rows)} total trades",
                "Market Title":            (
                    f"Live: {len(live_closed)} closed  ·  "
                    f"Research: {len(all_closed) - len(live_closed)} closed  ·  "
                    f"Pending: {pending_cnt}"
                ),
                "Recommendation":          f"Win Rate: {win_rate:.1f}%",
                "Result":                  f"{wins}W  /  {losses}L",
                "Profit/Loss ($)":         total_pnl,
                "Running Total P&L ($)":   total_pnl,
            }

            for c_idx, name in enumerate(DISPLAY_COLS, start=1):
                value          = summary.get(name, "")
                cell           = ws.cell(row=summary_r, column=c_idx, value=value)
                cell.fill      = _SUMMARY_FILL
                cell.font      = _SUMMARY_FONT
                cell.alignment = _LEFT if name == "Market Title" else _CENTER
                cell.border    = _BORDER
                if name in _MONEY_COLS and isinstance(value, (int, float)):
                    cell.number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'

        # ── Auto-size columns (based on actual content) ────────────────────
        for c_idx, name in enumerate(DISPLAY_COLS, start=1):
            col_letter = get_column_letter(c_idx)
            # padding of 4; cap at 55 so "Market Title" doesn't go crazy
            width = min(max(col_max[c_idx - 1] + 4, 9), 55)
            ws.column_dimensions[col_letter].width = width

        # ── Row heights ────────────────────────────────────────────────────
        ws.row_dimensions[1].height = 20
        for i in range(2, len(rows) + 2):
            ws.row_dimensions[i].height = 16

        wb.save(self._xlsx)
        log.info("excel_rebuilt", rows=len(rows), path=str(self._xlsx))

    # ── Stats ──────────────────────────────────────────────────────────────────

    def get_stats(self, *, today_only: bool = False) -> dict:
        """
        Compute summary statistics.
        today_only=True restricts to rows with today's UTC date.
        """
        rows = self._read()
        if today_only:
            today = date.today().isoformat()
            rows  = [r for r in rows if r.get("Date", "") == today]

        all_closed  = [r for r in rows if r.get("Result") in ("WIN", "LOSS")]
        live_closed = [r for r in all_closed if r.get("Mode", "").upper() == "LIVE"]
        wins        = sum(1 for r in all_closed if r.get("Result") == "WIN")
        pending_cnt = sum(1 for r in rows if r.get("Result") == "PENDING")
        win_rate    = wins / len(all_closed) * 100 if all_closed else 0.0

        total_pnl  = 0.0
        total_wgrd = 0.0
        pnl_rows: list[tuple[float, str, str]] = []

        for r in live_closed:
            try:
                pnl         = float(r.get("Profit/Loss ($)") or 0)
                total_pnl  += pnl
                total_wgrd += float(r.get("Amount Wagered ($)") or 0)
                pnl_rows.append((pnl, r.get("Market Title", ""), r.get("_ticker", "")))
            except ValueError:
                pass

        best  = max(pnl_rows, key=lambda x: x[0], default=(0.0, "—", ""))
        worst = min(pnl_rows, key=lambda x: x[0], default=(0.0, "—", ""))

        return {
            "total_trades":  len(rows),
            "closed_trades": len(all_closed),
            "live_trades":   len(live_closed),
            "wins":          wins,
            "losses":        len(all_closed) - wins,
            "pending":       pending_cnt,
            "win_rate":      round(win_rate, 1),
            "total_pnl":     round(total_pnl, 2),
            "total_wagered": round(total_wgrd, 2),
            "best_trade":    best,
            "worst_trade":   worst,
        }

    def get_pending_trades(self) -> list[dict]:
        """Return all rows currently marked PENDING."""
        rows = self._read()
        return [r for r in rows if r.get("Result") == "PENDING"]

    def get_xlsx_bytes(self) -> bytes:
        """Return the XLSX file contents for Discord attachment upload."""
        if not self._xlsx.exists():
            self._rebuild_excel()
        return self._xlsx.read_bytes()
